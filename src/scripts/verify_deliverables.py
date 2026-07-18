#!/usr/bin/env python3
"""Verify task block deliverables and write an audit report."""

from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = REPO_ROOT / "data"
AUDIT_DIR = DATA_DIR / "_audit"


@dataclass
class Check:
    block: str
    item: str
    status: str
    detail: str


def count_jsonl(path: Path) -> tuple[int, list[dict[str, Any]], list[str]]:
    """Count parseable JSONL rows and collect parse errors."""
    rows: list[dict[str, Any]] = []
    errors: list[str] = []
    for line_no, line in enumerate(path.read_text(encoding="utf-8").splitlines(), 1):
        if not line.strip():
            continue
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError as exc:
            errors.append(f"line {line_no}: {exc}")
    return len(rows), rows, errors


def count_xlsx_rows(path: Path) -> tuple[int, int]:
    """Count non-empty Excel data rows and columns across the active sheet."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    nonempty_rows = 0
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            nonempty_rows += 1
    return max(0, nonempty_rows - 1), ws.max_column


def add_threshold_check(
    checks: list[Check],
    block: str,
    item: str,
    actual: int,
    expected: int,
    unit: str = "条",
) -> None:
    """Append a PASS/FAIL threshold check."""
    status = "PASS" if actual >= expected else "FAIL"
    checks.append(Check(block, item, status, f"{actual} {unit}，验收下限 {expected} {unit}"))


def verify_block_01(checks: list[Check]) -> dict[str, Any]:
    """Verify block 01."""
    block = "block-01"
    base = DATA_DIR / "block-01-arbiteros-redteam-rewrite"
    xlsx_rows, xlsx_cols = count_xlsx_rows(base / "human_readable" / "arbiteros_cases_human_readable.xlsx")
    jsonl_count, rows, errors = count_jsonl(base / "gov_rewrite" / "arbiteros_cases_gov_rewrite.jsonl")
    expanded_count, _, expanded_errors = count_jsonl(
        base / "gov_rewrite" / "arbiteros_cases_gov_rewrite_expanded.jsonl"
    )
    md_count = len(list((base / "human_readable").glob("*.md")))
    source_index_size = (base / "arbiteros_case_source_index.md").stat().st_size

    add_threshold_check(checks, block, "人类可读 xlsx", xlsx_rows, 80)
    checks.append(Check(block, "人类可读 xlsx 列数", "PASS" if xlsx_cols >= 10 else "FAIL", f"{xlsx_cols} 列"))
    add_threshold_check(checks, block, "政务改写 JSONL", jsonl_count, 80)
    checks.append(Check(block, "政务改写 JSONL 可解析", "PASS" if not errors else "FAIL", f"{len(errors)} 个解析错误"))
    add_threshold_check(checks, block, "扩充版 JSONL", expanded_count, 247)
    checks.append(Check(block, "扩充版 JSONL 可解析", "PASS" if not expanded_errors else "FAIL", f"{len(expanded_errors)} 个解析错误"))
    add_threshold_check(checks, block, "人类可读 md", md_count, 80, "个")
    checks.append(Check(block, "源路径索引", "PASS" if source_index_size > 1000 else "FAIL", f"{source_index_size} bytes"))
    return {"xlsx_rows": xlsx_rows, "jsonl_rows": jsonl_count, "expanded_rows": expanded_count, "md_files": md_count}


def verify_block_02(checks: list[Check]) -> dict[str, Any]:
    """Verify block 02."""
    block = "block-02"
    base = DATA_DIR / "block-02-public-datasets-attack-patterns"
    xlsx_rows, xlsx_cols = count_xlsx_rows(base / "screening" / "public_benchmark_case_screening.xlsx")
    jsonl_count, rows, errors = count_jsonl(base / "gov_cases" / "public_patterns_to_gov_cases.jsonl")
    discarded_lines = len((base / "discarded" / "discarded_cases.md").read_text(encoding="utf-8").splitlines())
    duplicate_ids = [
        key
        for key, value in Counter(row.get("case_number") for row in rows).items()
        if key and value > 1
    ]

    add_threshold_check(checks, block, "公开筛选 xlsx", xlsx_rows, 1500)
    checks.append(Check(block, "公开筛选 xlsx 列数", "PASS" if xlsx_cols >= 10 else "FAIL", f"{xlsx_cols} 列"))
    add_threshold_check(checks, block, "政务改写 JSONL", jsonl_count, 27)
    checks.append(Check(block, "政务改写 JSONL 可解析", "PASS" if not errors else "FAIL", f"{len(errors)} 个解析错误"))
    checks.append(Check(block, "案例编号不重复", "PASS" if not duplicate_ids else "FAIL", f"{len(duplicate_ids)} 个重复编号"))
    add_threshold_check(checks, block, "排除案例说明", discarded_lines, 100, "行")
    return {"screening_rows": xlsx_rows, "jsonl_rows": jsonl_count, "discarded_lines": discarded_lines}


def verify_block_03(checks: list[Check]) -> dict[str, Any]:
    """Verify block 03."""
    block = "block-03"
    base = DATA_DIR / "block-03-gov-original-skills"
    skill_files = sorted((base / "skills").glob("gov-*-assistant/SKILL.md"))
    jsonl_count, rows, errors = count_jsonl(base / "cases" / "gov_original_cases.jsonl")

    add_threshold_check(checks, block, "OpenClaw Skill", len(skill_files), 5, "个")
    for skill_file in skill_files:
        text = skill_file.read_text(encoding="utf-8")
        status = "PASS" if len(text) >= 500 and "禁止" in text and "测试" in text else "FAIL"
        checks.append(Check(block, skill_file.parent.name, status, f"{len(text)} 字符"))
    add_threshold_check(checks, block, "原创案例 JSONL", jsonl_count, 53)
    checks.append(Check(block, "原创案例 JSONL 可解析", "PASS" if not errors else "FAIL", f"{len(errors)} 个解析错误"))
    return {"skills": len(skill_files), "jsonl_rows": jsonl_count}


def verify_block_04(checks: list[Check]) -> dict[str, Any]:
    """Verify block 04."""
    block = "block-04"
    base = DATA_DIR / "block-04-risk-grading-policy"
    matrix_rows, matrix_cols = count_xlsx_rows(base / "risk_level_matrix.xlsx")
    mapping_rows, mapping_cols = count_xlsx_rows(base / "case_to_policy_mapping.xlsx")
    rules_text = (base / "policy" / "gov_policy_rules.yaml").read_text(encoding="utf-8")
    semantic_text = (base / "policy" / "gov_semantic_rules_design.md").read_text(encoding="utf-8")

    add_threshold_check(checks, block, "风险等级矩阵", matrix_rows, 4)
    checks.append(Check(block, "风险等级矩阵列数", "PASS" if matrix_cols >= 4 else "FAIL", f"{matrix_cols} 列"))
    add_threshold_check(checks, block, "案例到规则映射", mapping_rows, 160)
    checks.append(Check(block, "案例映射列数", "PASS" if mapping_cols >= 10 else "FAIL", f"{mapping_cols} 列"))
    checks.append(Check(block, "策略规则 YAML", "PASS" if rules_text.count("GOV-") >= 12 else "FAIL", f"{rules_text.count('GOV-')} 个 GOV-* 标记"))
    checks.append(Check(block, "语义规则设计文档", "PASS" if len(semantic_text) > 500 else "FAIL", f"{len(semantic_text)} 字符；未验证运行时语义"))
    return {"matrix_rows": matrix_rows, "mapping_rows": mapping_rows, "policy_rule_markers": rules_text.count("GOV-")}


def verify_block_05(checks: list[Check]) -> dict[str, Any]:
    """Verify block 05."""
    block = "block-05"
    base = DATA_DIR / "block-05-arbiteros-batch-run"
    summary_path = base / "runs" / "20260712T025913.880037Z" / "summary.json"
    summary = json.loads(summary_path.read_text(encoding="utf-8"))
    index_rows, index_cols = count_xlsx_rows(base / "index" / "arbiteros_run_index.xlsx")
    output_dirs = [p for p in (base / "arbiteros_run_outputs").iterdir() if p.is_dir()]
    complete_output_dirs = 0
    for case_dir in output_dirs:
        if all((case_dir / name).is_dir() and any((case_dir / name).iterdir()) for name in ["results", "parsed", "raw"]):
            complete_output_dirs += 1

    checks.append(Check(block, "summary.json 总数", "PASS" if summary.get("total_cases") == 80 else "FAIL", str(summary.get("total_cases"))))
    checks.append(
        Check(
            block,
            "summary.json 归档计数一致性",
            "PASS" if summary.get("passed") == 74 and summary.get("failed") == 6 else "FAIL",
            f"归档字段 {summary.get('passed')}/{summary.get('total_cases')}；不评价 pass 判定语义",
        )
    )
    add_threshold_check(checks, block, "批跑索引 xlsx", index_rows, 80)
    checks.append(Check(block, "批跑索引列数", "PASS" if index_cols >= 10 else "FAIL", f"{index_cols} 列"))
    add_threshold_check(checks, block, "按 case_id 归档目录", len(output_dirs), 80, "个")
    add_threshold_check(checks, block, "完整 results/parsed/raw 目录", complete_output_dirs, 80, "个")
    return {
        "total_cases": summary.get("total_cases"),
        "passed": summary.get("passed"),
        "failed": summary.get("failed"),
        "pass_rate": summary.get("pass_rate"),
        "output_dirs": len(output_dirs),
        "complete_output_dirs": complete_output_dirs,
    }


def collect_skill_stats() -> list[dict[str, Any]]:
    """Collect case counts and run results by skill."""
    base = DATA_DIR / "arbiteros_standard_cases"
    id_to_skill = {path.stem: path.parent.name for path in base.glob("*/*.json")}
    summary = json.loads(
        (DATA_DIR / "block-05-arbiteros-batch-run" / "runs" / "20260712T025913.880037Z" / "summary.json").read_text(
            encoding="utf-8"
        )
    )
    stats: dict[str, Counter[str]] = defaultdict(Counter)
    for result in summary.get("results", []):
        skill = id_to_skill.get(result.get("id"), "UNKNOWN")
        stats[skill]["total"] += 1
        stats[skill][result.get("category", "unknown")] += 1
        stats[skill][result.get("status", "unknown")] += 1
    return [{"skill": skill, **dict(counter)} for skill, counter in sorted(stats.items())]


def write_reports(summary: dict[str, Any], checks: list[Check]) -> None:
    """Write Markdown and JSON audit reports."""
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    report_json = AUDIT_DIR / "structure_integrity_report.json"
    report_md = AUDIT_DIR / "structure_integrity_report.md"
    report_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    failed = [check for check in checks if check.status == "FAIL"]
    lines = [
        "# 1~5 号任务结构完整性验收报告",
        "",
        "> 自动生成。只检查文件存在性、数量门槛、字段和格式可解析性，不评价案例语义正确性，也不评价老师负责的 `system-design`。",
        f"> 生成时间：{summary['generated_at']}。",
        "",
        "## 总览",
        "",
        f"- 检查项：{len(checks)}",
        f"- 通过：{len(checks) - len(failed)}",
        f"- 失败：{len(failed)}",
        f"- 结构完整性结论：{'✅ 通过' if not failed else '❌ 需修复'}",
        "",
        "## 分块结果",
        "",
        "| 任务块 | 检查项 | 状态 | 说明 |",
        "|---|---|---|---|",
    ]
    for check in checks:
        mark = "✅" if check.status == "PASS" else "❌"
        lines.append(f"| {check.block} | {check.item} | {mark} {check.status} | {check.detail} |")

    lines.extend(
        [
            "",
            "## 按 Skill 批跑统计",
            "",
            "| Skill | 总数 | Safe | Unsafe | Pass | Fail |",
            "|---|---:|---:|---:|---:|---:|",
        ]
    )
    for row in summary["skill_stats"]:
        lines.append(
            f"| {row['skill']} | {row.get('total', 0)} | {row.get('safe', 0)} | "
            f"{row.get('unsafe', 0)} | {row.get('pass', 0)} | {row.get('fail', 0)} |"
        )

    report_md.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--write", action="store_true", help="write reports to data/_audit")
    args = parser.parse_args()

    checks: list[Check] = []
    block_summary = {
        "block_01": verify_block_01(checks),
        "block_02": verify_block_02(checks),
        "block_03": verify_block_03(checks),
        "block_04": verify_block_04(checks),
        "block_05": verify_block_05(checks),
    }
    summary = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "scope": "block-01..block-05 only; system-design excluded",
        "checks": [asdict(check) for check in checks],
        "block_summary": block_summary,
        "skill_stats": collect_skill_stats(),
    }
    failed = [check for check in checks if check.status == "FAIL"]
    for check in checks:
        mark = "PASS" if check.status == "PASS" else "FAIL"
        print(f"[{mark}] {check.block} / {check.item}: {check.detail}")
    print(f"\nSUMMARY: {len(checks) - len(failed)} passed, {len(failed)} failed out of {len(checks)}")
    if args.write:
        write_reports(summary, checks)
        print(f"Wrote {AUDIT_DIR / 'structure_integrity_report.md'}")
        print(f"Wrote {AUDIT_DIR / 'structure_integrity_report.json'}")
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
